from openpyxl import Workbook
from .models import *
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from datetime import date, timedelta, datetime
from business.models import Loan
from django.db.models import Sum	



@login_required
def sales_report_export_excel(request, **kwargs):
	sales = kwargs.get('sales')
	
	file_name = f"General_Sales_Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook()   
	worksheet = workbook.active
	worksheet.title = f'General Sales Report'

	 # Define the titles for columns
	columns = ['#','BUSINESS BRANCH','ORDER NO.','PRODUCT','QUANTITY','TOTAL','CVP','PAID','DISCOUNT','TAX', 'CREATED']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(columns, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title

	for sale in sales:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			sale.business.name,
			sale.order_no,
			sale.inventory.product.name,
			sale.quantity,
			sale.total,
			sale.profit,
			sale.amount_paid,
			f'{sale.discount} {sale.discount_unit}',
			f'{sale.tax} {sale.tax_unit}',
			sale.created_at.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = worksheet.cell(row=row_num, column=col_num)
			cell.value = cell_value

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response  	



@login_required
def inventory_report_export_excel(request, **kwargs):
	inventories = kwargs.get('inventories')

	file_name = f"General_Inventory_Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook()   
	worksheet = workbook.active
	worksheet.title = f'General Inventory Report'

	 # Define the titles for columns
	columns = ['#','BUSINESS BRANCH','PRODUCT','QUANTITY','AVAILABLE','DAMAGE','PRODUCT COST','SELL PRICE','COGS', 'CREATED']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(columns, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title

	for inventory in inventories:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			inventory.business.name,
			inventory.product.name,
			inventory.quantity,
			inventory.remain-inventory.damage,
			inventory.damage,			
			inventory.product_cost,
			inventory.product.sell_price,
			inventory.cogs,
			inventory.created_at.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = worksheet.cell(row=row_num, column=col_num)
			cell.value = cell_value

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 


@login_required
def procurement_report_export_excel(request, **kwargs):
	business = kwargs.get('business')
	local_purchases = kwargs.get('local_purchases')
	purchases = kwargs.get('purchases') 

	file_name = f"General_Procurement_Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook() 

	# Local Purchases order
	ws1 = workbook.active
	ws1.title = f'General Local Purchases Report'
	ws1.sheet_properties.tabColor = "007bff"

	 # Define the titles for columns
	columns = ['#','BUSINESS BRANCH','LPO','SUPPLIER','DELIVERY','PREPARED BY','TOTAL','CREATED','AUTHORIZED']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = column_title

	for local in local_purchases:  	
		# Define the data for each cell in the row 
		for auth in local.positive_authorize_list:
			authorized = auth.created_at                                          
                                        
		row = [
			row_num,
			local.business.name,
			local.lpo_no,
			local.supplier.name,
			local.delivery,
			local.employee.user_employee.position,			
			local.total,
			local.created_at.strftime('%d-%m-%Y'),
			authorized.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = ws1.cell(row=row_num, column=col_num)
			cell.value = cell_value

	# Purchases order
	ws2 = workbook.create_sheet()
	ws2.title = f'General Purchases Report'
	ws2.sheet_properties.tabColor = "dc3545"

	 # Define the titles for columns
	columns = ['#','BUSINESS BRANCH','PO','SUPPLIER','DELIVERY','PREPARED BY','TOTAL','CREATED','AUTHORIZED']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(columns, 1):
		cell = ws2.cell(row=row_num, column=col_num)
		cell.value = column_title

	for local in purchases:  	
		# Define the data for each cell in the row 
		for auth in local.positive_authorize_list:
			authorized = auth.created_at   
		row = [
			row_num,
			local.business.name,
			local.po_no,
			local.supplier.name,
			local.delivery,
			local.employee.user_employee.position,			
			local.total,
			local.created_at.strftime('%d-%m-%Y'),
			authorized.strftime('%d-%m-%Y'),	
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = ws2.cell(row=row_num, column=col_num)
			cell.value = cell_value	



	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 


@login_required
def opex_report_export_excel(request, **kwargs):
	expenses = kwargs.get('expense_objects')
	total = expenses.aggregate(Sum('cost'))['cost__sum']
	print(total)

	file_name = f"General_OPEX_Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook() 

	# Local Purchases order
	ws1 = workbook.active
	ws1.title = f'General OPEX Report'

	 # Define the titles for columns
	columns = ['#','BUSINESS BRANCH','NAME','CATEGORY','BANK','BANK ACCOUNT NO','COST','DATE',]
	row_num = 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	for expense in expenses:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			expense.business.name,
			expense.name,
			expense.category,
			expense.bank.branch,
			expense.account,
			expense.cost,			
			expense.date.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = ws1.cell(row=row_num, column=col_num)
			cell.value = cell_value	

	 # Define the titles for columns
	columns = ['#','TOTAL','','','','',total,'',]
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 	


@login_required
def payroll_report_export_excel(request, **kwargs):
	business = kwargs.get('business')
	takehome = kwargs.get('takehome')
	takehome_total = kwargs.get('takehome_total')
	nssf = kwargs.get('nssf')
	paye = kwargs.get('paye')
	loan_board = kwargs.get('loan_board')
	sdl = kwargs.get('sdl')
	bonus = kwargs.get('bonus')
	overtime = kwargs.get('overtime')
	deduction = kwargs.get('deduction')
	salaries_total = kwargs.get('salaries_total')
	file_name = f"Payroll_General_Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook() 

	# Local Purchases order
	ws1 = workbook.active
	ws1.title = f'Payroll General Report'

	 # Define the titles for columns
	columns = ['#','BUSINESS BRANCH','EMPLOYEE ID','NAME','POSITION','DEPARTMENT','SALARY','TAKEHOME','PAYE','NSSF','HESLB','BONUS','OVERTIME','DEDUCTION','SDL','CREATED']
	row_num = 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	for worker in takehome:  	
		# Define the data for each cell in the row 
		loan = Loan.objects.filter(employee=worker.payroll.employee, business=business).first()

		row = [
			row_num,
			worker.payroll.business.name,
			worker.payroll.employee.id_no,
			worker.payroll.employee.full_name,
			worker.payroll.employee.position,
			worker.payroll.employee.department.name,
			worker.payroll.employee.salary,
			worker.salary,	
			worker.payroll.paye_amount,
			worker.payroll.nssf_amount,
			worker.payroll.heslb_amount,
			worker.payroll.bonus,
			worker.payroll.overtime,
			worker.payroll.deduction,
			worker.payroll.sdl_amount,	
			worker.payroll.created_at.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = ws1.cell(row=row_num, column=col_num)
			cell.value = cell_value	


	columns = ['#','TOTAL','','','','',salaries_total,takehome_total,paye,nssf,loan_board,bonus,overtime,deduction,sdl,'']
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 


@login_required
def stock_report_export_excel(request, **kwargs):
	business = kwargs.get('business')
	stocks = kwargs.get('stocks')

	file_name = f"General_Stock_Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook()   
	worksheet = workbook.active
	worksheet.title = f'General Stock Report'

	 # Define the titles for row
	row = ['#','BUSINESS BRANCH','PRODUCT NAME','QUANTITY','PRODUCT COST','TYPE','DATE']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title

	for stock in stocks:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			stock.business.name,
			stock.product.name,
			stock.quantity,
			stock.cost,
			stock.stock_type,			
			stock.stock_date.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = worksheet.cell(row=row_num, column=col_num)
			cell.value = cell_value

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 	