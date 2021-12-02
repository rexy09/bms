from openpyxl import Workbook
from administration.models import *
from .models import *
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from datetime import date, timedelta, datetime
from decimal import Decimal


@login_required
def sales_report_export_excel(request, **kwargs):
	business = kwargs.get('business')
	sales = kwargs.get('sales')

	file_name = f"{business.name}_Sales_Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook()   
	worksheet = workbook.active
	worksheet.title = f'{business.name} Sales Report'

	 # Define the titles for row
	row = ['#','ORDER NO.','PRODUCT','QUANTITY','TOTAL','CVP','PAID','DISCOUNT','TAX', 'CREATED']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title

	for sale in sales:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			sale.order_no,
			sale.inventory.product.name,
			sale.quantity,
			sale.total,
			sale.profit,
			sale.amount_paid,
			f'{sale.discount} {sale.discount_unit}',
			f'{sale.tax} {sale.tax_unit}',
			sale.created_at.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = worksheet.cell(row=row_num, column=col_num)
			cell.value = cell_value

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response  		


@login_required
def inventory_report_export_excel(request, **kwargs):
	business = kwargs.get('business')
	inventories = kwargs.get('inventories')

	file_name = f"{business.name}_Inventory_Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook()   
	worksheet = workbook.active
	worksheet.title = f'{business.name} Inventory Report'

	 # Define the titles for row
	row = ['#','PRODUCT','QUANTITY','AVAILABLE','DAMAGE','PRODUCT COST','SELL PRICE','COGS', 'CREATED']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title

	for inventory in inventories:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			inventory.product.name,
			inventory.quantity,
			inventory.remain-inventory.damage,
			inventory.damage,			
			inventory.product_cost,
			inventory.product.sell_price,
			inventory.cogs,
			inventory.created_at.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = worksheet.cell(row=row_num, column=col_num)
			cell.value = cell_value

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 


@login_required
def procurement_report_export_excel(request, **kwargs):
	business = kwargs.get('business')
	local_purchases = kwargs.get('local_purchases')
	purchases = kwargs.get('purchases') 

	file_name = f"{business.name}_Procurement_Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook() 

	# Local Purchases order
	ws1 = workbook.active
	ws1.title = f'{business.name} Local Purchases Report'
	ws1.sheet_properties.tabColor = "007bff"

	 # Define the titles for row
	row = ['#','LPO','SUPPLIER','DELIVERY','PREPARED BY','TOTAL','CREATED','AUTHORIZED']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = column_title

	for local in local_purchases:  	
		# Define the data for each cell in the row 
		for auth in local.positive_authorize_list:
			authorized = auth.created_at 
		row = [
			row_num,
			local.lpo_no,
			local.supplier.name,
			local.delivery,
			local.employee.user_employee.position,			
			local.total,
			local.created_at.strftime('%d-%m-%Y'),
			authorized.strftime('%d-%m-%Y'),	
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = ws1.cell(row=row_num, column=col_num)
			cell.value = cell_value

	# Purchases order
	ws2 = workbook.create_sheet()
	ws2.title = f'{business.name} Purchases Report'
	ws2.sheet_properties.tabColor = "dc3545"

	 # Define the titles for row
	row = ['#','PO','SUPPLIER','DELIVERY','PREPARED BY','TOTAL','CREATED','AUTHORIZED']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row, 1):
		cell = ws2.cell(row=row_num, column=col_num)
		cell.value = column_title

	for local in purchases:  	
		# Define the data for each cell in the row 
		for auth in local.positive_authorize_list:
			authorized = auth.created_at 
		row = [
			row_num,
			local.po_no,
			local.supplier.name,
			local.delivery,
			local.employee.user_employee.position,			
			local.total,
			local.created_at.strftime('%d-%m-%Y'),
			authorized.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = ws2.cell(row=row_num, column=col_num)
			cell.value = cell_value	


	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 


@login_required
def opex_report_export_excel(request, **kwargs):
	business = kwargs.get('business')
	expenses = kwargs.get('expense_objects')

	file_name = f"{business.name}_OPEX_Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook() 

	# Local Purchases order
	ws1 = workbook.active
	ws1.title = f'{business.name} OPEX Report'

	 # Define the titles for columns
	columns = ['#','NAME','CATEGORY','BANK','BANK ACCOUNT NO','COST','DATE',]
	row_num = 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	for expense in expenses:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			expense.name,
			expense.category,
			expense.bank.branch,
			expense.account,
			expense.cost,			
			expense.date.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = ws1.cell(row=row_num, column=col_num)
			cell.value = cell_value	



	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 


@login_required
def payroll_report_export_excel(request, **kwargs):
	business = kwargs.get('business')
	takehome = kwargs.get('takehome')
	takehome_total = kwargs.get('takehome_total')
	nssf = kwargs.get('nssf')
	paye = kwargs.get('paye')
	loan_board = kwargs.get('loan_board')
	sdl = kwargs.get('sdl')
	bonus = kwargs.get('bonus')
	overtime = kwargs.get('overtime')
	deduction = kwargs.get('deduction')
	salaries_total = kwargs.get('salaries_total')

	file_name = f"{business.name}_Payroll_Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook() 

	# Local Purchases order
	ws1 = workbook.active
	ws1.title = f'{business.name} Payroll Report'


	 # Define the titles for columns
	columns = [business.name.upper()]
	row_num = 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	 # Define the titles for columns
	columns = [" "]
	row_num = 2

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	 # Define the titles for columns
	columns = ['#','EMPLOYEE ID','NAME','POSITION','DEPARTMENT','SALARY','TAKEHOME','PAYE','NSSF','HESLB','BONUS','OVERTIME','DEDUCTION','SDL','CREATED']
	row_num = 3

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	for worker in takehome:  	
		# Define the data for each cell in the row 
		loan = Loan.objects.filter(employee=worker.payroll.employee, business=business).first()

		row = [
			row_num,
			worker.payroll.employee.id_no,
			worker.payroll.employee.full_name,
			worker.payroll.employee.position,
			worker.payroll.employee.department.name,
			worker.payroll.employee.salary,
			worker.salary,	
			worker.payroll.paye_amount,
			worker.payroll.nssf_amount,
			worker.payroll.heslb_amount,
			worker.payroll.bonus,
			worker.payroll.overtime,
			worker.payroll.deduction,
			worker.payroll.sdl_amount,	
			worker.payroll.created_at.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = ws1.cell(row=row_num, column=col_num)
			cell.value = cell_value	


	columns = ['#','TOTAL','','','',salaries_total,takehome_total,paye,nssf,loan_board,bonus,overtime,deduction,sdl,'']
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value				


	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 
	



@login_required
def trial_balance_export_excel(request, **kwargs):
	id = kwargs.get('id')
	business = Business.objects.filter(id=id).first()
	fr = request.GET.get('fr')
	to = request.GET.get('to')	
	business = Business.objects.filter(id=id).first()
	nssf_funds = Payroll.objects.filter(nssf=True, created_at__date__gte=fr, created_at__date__lte=to).aggregate(Sum('nssf_amount'))['nssf_amount__sum']
	loan_board_funds = Payroll.objects.filter(heslb=True, created_at__date__gte=fr, created_at__date__lte=to).aggregate(Sum('heslb_amount'))['heslb_amount__sum']
	paye_total = Payroll.objects.filter(paye=True, created_at__date__gte=fr, created_at__date__lte=to).aggregate(Sum('paye_amount'))['paye_amount__sum']	
	takehome_total = Takehome.objects.all().aggregate(Sum('salary'))['salary__sum']
	sdl = Payroll.objects.filter(created_at__date__gte=fr, created_at__date__lte=to).aggregate(Sum('sdl_amount'))['sdl_amount__sum']   		
	taxes = Tax.objects.filter(created_at__date__gte=fr, created_at__date__lte=to).aggregate(Sum('remain'))['remain__sum']
	interest_remaining = Interest.objects.filter(created_at__date__gte=fr, created_at__date__lte=to).aggregate(Sum('remaining'))['remaining__sum']
	interest_repayment = Interest.objects.filter(created_at__date__gte=fr, created_at__date__lte=to).aggregate(Sum('repayment'))['repayment__sum']
	sales = Sale.objects.filter(created_at__date__gte=fr, created_at__date__lte=to).all().aggregate(total_paid=Sum('amount_paid'))

	inventory_qs = Inventory.objects.filter(created_at__date__gte=fr, created_at__date__lte=to).annotate(available=F('remain')-F('damage'))			
	inventories = inventory_qs.filter(available__gt=0, ).order_by('-pk')

	# Start COGS
	purchases = Inventory.objects.filter(created_at__date__gte=fr, created_at__date__lte=to,).aggregate(total_cost=Sum(F('quantity')*F('product_cost'),output_field=FloatField()))['total_cost']		
	opening = Stock.objects.filter(stock_date__gte=fr, stock_date__lte=to, stock_type='Opening Stock').aggregate(Sum('cost'))['cost__sum']
	closing = Stock.objects.filter(stock_date__gte=fr, stock_date__lte=to, stock_type='Closing Stock').aggregate(Sum('cost'))['cost__sum']
	if purchases is None:
		purchases = 0
	if opening is None:
		opening = 0
	if closing is None:
		closing = 0		
	cogs = opening + Decimal(purchases) - closing
	# End COGS	
	
	expenses = Expense.objects.filter(created_at__date__gte=fr, created_at__date__lte=to).aggregate(Sum('cost'))['cost__sum']
	liabilities = Liability.objects.filter(created_at__date__gte=fr, created_at__date__lte=to).aggregate(Sum('cost'))['cost__sum']
	fixed_assets = AccountsFixedAsset.objects.filter(created_at__date__gte=fr, created_at__date__lte=to).aggregate(Sum('value'))['value__sum']
	current_assets = AccountsCurrentAsset.objects.filter(created_at__date__gte=fr, created_at__date__lte=to).aggregate(Sum('cost'))['cost__sum']

	sales = sales['total_paid']

	if fixed_assets is None:
		fixed_assets = 0
	if current_assets is None:
		current_assets = 0
	if interest_remaining is None:
		interest_remaining = 0	
	if interest_repayment is None:
		interest_repayment = 0		
	
	interests = interest_repayment - interest_remaining		

	assets = fixed_assets + current_assets		


	if takehome_total is None:
		takehome_total = 0
	if sdl is None:
		sdl = 0	
	if taxes is None:
		taxes = 0
	if interests is None:
		interests = 0					
	if liabilities is None:
		liabilities = 0
	if assets is None:
		assets = 0
	if cogs is None:
		cogs = 0
	if paye_total is None:
		paye_total = 0	
	if expenses is None:
		expenses = 0
	if sales is None:
		sales = 0
	if nssf_funds is None:
		nssf_funds = 0
	if loan_board_funds is None:
		loan_board_funds = 0
	if paye_total is None:
		paye_total = 0			
	if cogs is None:
		cogs = 0



	total_debit = int(nssf_funds) + int(paye_total) + int(loan_board_funds) + int(takehome_total) + int(sdl) + int(interests) + int(taxes) + int(expenses) + int(cogs) + int(assets)
	total_credit = int(sales) + int(liabilities)

	file_name = f"Clask_Trial_Balance_Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook()   
	worksheet = workbook.active
	worksheet.title = f'Clask Trial Balance Report'

	 # Define the titles for row
	row1 = ['DESCRIPTION', 'DEBIT', 'CREDIT']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row1, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title

	row3 = ['Assets', int(assets),'0']
	row_num = 3

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row3, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row4 = ['Liabilities','0', int(liabilities)]
	row_num = 4

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row4, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row5 = ['Salaries', int(takehome_total), '0']
	row_num = 5

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row5, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row6 = ['Taxes', int(taxes), '0']
	row_num = 6

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row6, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row7 = ['Interests', int(interests), '0']
	row_num = 7

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row7, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row8 = ['Paye', int(paye_total), '0']
	row_num = 8

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row8, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row9 = ['Sdl', int(sdl), '0']
	row_num = 9

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row9, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row10 = ['Expenses', int(expenses), '0']
	row_num = 10

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row10, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row11 = ['Cogs', int(cogs), '0']
	row_num = 11

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row11, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row12 = ['Sales', '0', int(sales)]
	row_num = 12

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row12, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row13 = ['Heslb', int(loan_board_funds), '0']
	row_num = 13

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row13, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title
	row14 = ['Nssf', int(nssf_funds), '0']
	row_num = 14

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row14, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title

	row16 = ['Total', int(total_debit), int(total_credit)]
	row_num = 15

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row16, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title		



	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response	


def balance_sheet_export_excel(request, **kwargs):
	business = kwargs.get('business')
	assets = kwargs.get('assets')
	liabilities = kwargs.get('liabilities')
	equity = kwargs.get('equity')
	assets_total = kwargs.get('assets_total'),
	liabilities_total = kwargs.get('liabilities_total'),
	equity_total = kwargs.get('equity_total'),
	liabilities = kwargs.get('liabilities'),
	fixed_assets = kwargs.get('fixed_assets'),
	equities = kwargs.get('equities'),
	liability_equity = kwargs.get('liability_equity'),
	current_assets = kwargs.get('current_assets'),	

	file_name = f"Clask_Balance_Sheet_Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook() 

	# Local Purchases order
	ws1 = workbook.active
	ws1.title = f'Clask Balance Sheet Report'

	 # Define the titles for columns


	columns = ['PARTICULARS','AMOUNT']
	row_num = 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	row = ['Current Assets']
	row_num = 2

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(row, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	for assets in current_assets:
		for asset in assets:	  	
			# Define the data for each cell in the row 
			row = [
				asset.name,
				asset.cost,			
			]
			row_num += 1

			# Assign the data for each cell of the row 
			for col_num, cell_value in enumerate(row, 1):
				cell = ws1.cell(row=row_num, column=col_num)
				cell.value = cell_value

	for assets in fixed_assets:
		for asset in assets:	   	
			# Define the data for each cell in the row 
			row = [
				asset.name,
				asset.value,			
			]
			row_num += 1

			# Assign the data for each cell of the row 
			for col_num, cell_value in enumerate(row, 1):
				cell = ws1.cell(row=row_num, column=col_num)
				cell.value = cell_value

	row = ['Total Assets', assets_total[0]]
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(row, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value					

	row = ['Liabilities']
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(row, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value


	for values in liabilities:
		for liability in values:	   	
			# Define the data for each cell in the row 
			row = [
				liability.name,
				liability.cost,			
			]
			row_num += 1

			# Assign the data for each cell of the row 
			for col_num, cell_value in enumerate(row, 1):
				cell = ws1.cell(row=row_num, column=col_num)
				cell.value = cell_value		


	row = ['Total Liabilities', liabilities_total[0]]
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(row, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value		

	row = ['Equities']
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(row, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value


	for values in equities:
		for equity in values:	   	
			# Define the data for each cell in the row 
			row = [
				equity.name,
				equity.value,			
			]
			row_num += 1

			# Assign the data for each cell of the row 
			for col_num, cell_value in enumerate(row, 1):
				cell = ws1.cell(row=row_num, column=col_num)
				cell.value = cell_value	



	row = ['Total Equities', equity_total[0]]
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(row, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value


	row = [f'Liabilities plus Equities {liability_equity[0]}', f'Total Assets {assets_total[0]}']
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(row, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value




	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 

def income_statement_export_excel(request, **kwargs):
	summary = kwargs.get('summary')	
	business = kwargs.get('business')
	operating_revenue = summary['total_paid']
	operational_expense = summary['opex']
	tax_interest = summary['tax_interest']
	cogs = summary['cogs']
	total_income = summary['total_paid']
	net_income = summary['net_income']
	ebitda = summary['ebitda']
	ebit = summary['ebit']
	expenses = summary['expenses']
	da = summary['da']
	salaries = summary['salaries']
	gross_profit = summary['gross_profit']

	advertising = summary['advertising'] 
	contractors = summary['contractors'] 
	rent = summary['rent'] 
	travel = summary['travel'] 
	financial_charge = summary['financial_charge'] 
	utilities = summary['utilities']  
	services = summary['services']  
	other_expenses = summary['other_expenses']  
	education = summary['education']  

	file_name = f"clask_Income_Statement_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook() 

	# Local Purchases order
	ws1 = workbook.active
	ws1.title = f'clask Income Statement'

	 # Define the titles for columns
	columns = ['PARTICULARS', 'AMOUNT']
	row_num = 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	# Define the titles for columns
	columns = ['Operating Revenue', operating_revenue]
	row_num = 2

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	# Define the titles for columns
	columns = ['TOTAL INCOME', total_income]
	row_num = 3

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	 # Define the titles for columns
	columns = ['Cogs', round(cogs,2)]
	row_num = 4

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value		


	 # Define the titles for columns
	columns = ['GROSS PROFIT', round(gross_profit,2)]
	row_num = 5

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	 # Define the titles for columns
	columns = ['Advertising', advertising]
	row_num = 6

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	 # Define the titles for columns
	columns = ['Contractors', contractors]
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	 # Define the titles for columns
	columns = ['Rent or Lease', rent]
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	 # Define the titles for columns
	columns = ['Travel', travel]
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	 # Define the titles for columns
	columns = ['Financial Charge', financial_charge]
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	 # Define the titles for columns
	columns = ['Education and Training', education]
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	 # Define the titles for columns
	columns = ['Utilities', utilities]
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value					

	 # Define the titles for columns
	columns = ['Professional Services', services]
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	 # Define the titles for columns
	columns = ['Other Expenses', other_expenses]
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value			


	 # Define the titles for columns
	columns = ['Salaries Expenses', salaries]
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	 # Define the titles for columns
	columns = ['Depreciation & Amortization', da]
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	
		

	 # Define the titles for columns
	columns = ['OPERATIONAL EXPENSE', operational_expense]
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	 # Define the titles for columns
	columns = ['EBIT', ebit]
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	#  # Define the titles for columns
	# columns = ['EBITDA', ebitda]
	# row_num = 8

	# # Assign the data for each cell of the row 
	# for col_num, cell_value in enumerate(columns, 1):
	# 	cell = ws1.cell(row=row_num, column=col_num)
	# 	cell.value = cell_value	



	 # Define the titles for columns
	columns = ['Tax & Interest', tax_interest]
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	


	 # Define the titles for columns
	columns = ['NET INCOME', net_income]
	row_num += 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response	

def cashbook_export_excel(request, **kwargs):
	business = kwargs.get('business')
	cashflow = kwargs.get('cashflow')

	file_name = f"{business.name}_Cashbook_Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook() 

	# Local Purchases order
	ws1 = workbook.active
	ws1.title = f'{business.name} Cashbook Report'

	 # Define the titles for columns
	columns = ['#','DATE CREATED','DESCRIPTION','DEBIT','CREDIT','BALANCE']
	row_num = 1

	# Assign the data for each cell of the row 
	for col_num, cell_value in enumerate(columns, 1):
		cell = ws1.cell(row=row_num, column=col_num)
		cell.value = cell_value	

	for cashbook in cashflow:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			cashbook.created_at.strftime('%d-%m-%Y'),
			cashbook.description,
			cashbook.debit,			
			cashbook.credit,
			cashbook.balance,
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = ws1.cell(row=row_num, column=col_num)
			cell.value = cell_value	



	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response


@login_required
def stock_report_export_excel(request, **kwargs):
	business = kwargs.get('business')
	stocks = kwargs.get('stocks')

	file_name = f"{business.name}_Stock_Report_{datetime.today().strftime('%d_%m_%Y_%H:%M:%S')}.xlsx" 

	workbook = Workbook()   
	worksheet = workbook.active
	worksheet.title = f'{business.name} Stock Report'

	 # Define the titles for row
	row = ['#','PRODUCT NAME','QUANTITY','PRODUCT COST','TYPE','DATE']
	row_num = 1

	# Assign the titles for each cell of the header
	for col_num, column_title in enumerate(row, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title

	for stock in stocks:  	
		# Define the data for each cell in the row 
		row = [
			row_num,
			stock.product.name,
			stock.quantity,
			stock.cost,
			stock.stock_type,			
			stock.stock_date.strftime('%d-%m-%Y'),
		]
		row_num += 1

		# Assign the data for each cell of the row 
		for col_num, cell_value in enumerate(row, 1):
			cell = worksheet.cell(row=row_num, column=col_num)
			cell.value = cell_value

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename={file_name}'
	workbook.save(response) 
	return response 